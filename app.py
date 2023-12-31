from flask import Flask, render_template, request, redirect, url_for, session, Response, jsonify
import mysql.connector
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
from datetime import datetime
import plotly.express as px
import pandas as pd
import json
import re
from reportlab.pdfgen import canvas
from flask import Response
import os
# Todo lo de arriba se instala
import locale


app = Flask(__name__)
app.secret_key = "your_secret_key"
db = mysql.connector.connect(
    host="localhost", user="root", database="gnominane"
)
cursor = db.cursor()

# Pagina principal


@app.route("/")
def layout():
    return render_template('layout.html')


# Pagina de inicio de sesión
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        cursor = db.cursor()
        query = "SELECT * FROM usuario WHERE nombre = %s"
        cursor.execute(query, (username,))
        user = cursor.fetchone()

        if user and user[2] == password:
            session["username"] = user[1]
            return redirect(url_for("index"))
        else:
            error = "Usuario o contraseña incorrectos"

    return render_template('login.html', error=error if 'error' in locals() else None)


# Función de registro
@app.route('/registro', methods=['GET', 'POST'])
def registro():
    error_password = None
    error_correo = None
    sucefull = None

    if request.method == 'POST':
        Nombre = request.form['Nombre']
        Contraseña = request.form['Contraseña']
        Correo = request.form['Correo']

        # Validar longitud mínima y máxima para la contraseña
        if len(Contraseña) < 8:
            error_password = "La contraseña debe tener al menos 8 caracteres."
        if len(Contraseña) > 10:
            error_password = "La contraseña no puede tener más de 20 caracteres."

        # Verificar si el correo ya está registrado
        cursor = db.cursor()
        cursor.execute("SELECT * FROM usuario WHERE correo = %s", (Correo,))
        existing_user = cursor.fetchone()

        if existing_user:
            error_correo = "Este correo electrónico ya está registrado."

        # Validar que la contraseña contenga al menos un número, una mayúscula y una minúscula
        if not re.search(r'[0-9]', Contraseña) or not re.search(r'[A-Z]', Contraseña) or not re.search(r'[a-z]', Contraseña):
            error_password = "La contraseña debe contener al menos un número, una mayúscula y una minúscula."

        if not error_password and not error_correo:
            # Inserta el nuevo usuario en la base de datos
            cursor.execute(
                "INSERT INTO usuario (Nombre, Contraseña, correo) VALUES (%s, %s, %s)", (Nombre, Contraseña, Correo))
            db.commit()
            sucefull = "Registro exitoso. Ahora puedes iniciar sesión."

    return render_template('login.html', error_password=error_password, error_correo=error_correo, sucefull=sucefull)


# Lista de trabajadores
@app.route("/trabajadores")
def trabajadores():
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT e.*, d.Nombre AS Nombre_Departamento, c.Nombre AS Nombre_Cargo
        FROM empleados e
        JOIN departamento d ON e.Departamento = d.ID
        JOIN cargos c ON e.Cargo_ID = c.ID
        WHERE e.Estado = 'Activo'
    """)

    empleados_activos = cursor.fetchall()
    cursor.close()

    return render_template("trabajadores.html", empleados=empleados_activos)


# lista de cargos
@app.route("/cargos", methods=['GET', 'POST'])
def cargos():
    if "username" not in session:
        return redirect(url_for("login"))

    if request.method == 'POST':
        departamento_id = request.form.get('departamento')

        # Consulta SQL para obtener la lista de cargos según el departamento seleccionado
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM cargos WHERE Departamento_ID = %s
        """, (departamento_id,))
        lista_cargos = cursor.fetchall()
        cursor.close()

        # Consulta SQL para obtener la lista de departamentos
        cursor = db.cursor(dictionary=True)
        cursor.execute("""
            SELECT * FROM departamento
        """)
        lista_departamentos = cursor.fetchall()
        cursor.close()

        return render_template("cargos.html", departamentos=lista_departamentos, cargos=lista_cargos)

    # Consulta SQL para obtener la lista de departamentos
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT * FROM departamento
    """)
    lista_departamentos = cursor.fetchall()
    cursor.close()

    return render_template("cargos.html", departamentos=lista_departamentos, cargos=[])


# Pagina principal de Nomina
@app.route("/index")
def index():
    if "username" not in session:
        return redirect(url_for("login"))
    return render_template("index.html")


@app.route("/usuarios")
def usuarios():
    if "username" not in session:
        return redirect(url_for("login"))
    return render_template("usuarios.html")


# Lista de departamentos
@app.route("/departamentos", methods=["GET", "POST"])
def listar_departamentos():
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM departamento WHERE estado='Activo'")
    departamentos = cursor.fetchall()
    cursor.close()

    return render_template("departamentos.html", departamentos=departamentos)


# Lista de departamentos inactivos
@app.route("/departamentos_Inactivos", methods=["GET", "POST"])
def departamentos_Inactivos():
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM departamento WHERE estado='Inactivo'")
    departamentos = cursor.fetchall()
    cursor.close()

    return render_template("departamentos_Inactivos.html", departamentos=departamentos)


# Función de desactivar departamento
@app.route("/desactivar_departamento/<int:id_departamento>")
def desactivar_departamento(id_departamento):
    if "username" not in session:
        return redirect(url_for("login"))

    # Actualiza el estado del departamento a "Inactivo" en tu base de datos
    cursor = db.cursor()
    cursor.execute(
        "UPDATE departamento SET Estado = 'Inactivo' WHERE ID = %s", (id_departamento,))
    db.commit()
    cursor.close()

    # Redirige de nuevo a la lista de departamentos después de desactivar
    return redirect(url_for("listar_departamentos"))


# Función de reactivar departamento
@app.route("/reactivar_departamento/<int:id_departamento>")
def reactivar_departamento(id_departamento):
    if "username" not in session:
        return redirect(url_for("login"))

    # Actualiza el estado del departamento a "Inactivo" en tu base de datos
    cursor = db.cursor()
    cursor.execute(
        "UPDATE departamento SET Estado = 'Activo' WHERE ID = %s", (id_departamento,))
    db.commit()
    cursor.close()

    # Redirige de nuevo a la lista de departamentos después de desactivar
    return redirect(url_for("listar_departamentos"))


@app.route("/agregar_departamento", methods=["GET", "POST"])
def agregar_departamento():
    if "username" not in session:
        return redirect(url_for("login"))

    departamentos = []  # Inicializar la variable departamentos

    if request.method == "POST":
        nombre_departamento = request.form["nombre"]
        Fecha_Creacion = request.form["Fecha"]

        cursor = db.cursor()
        cursor.execute(
            "SELECT ID FROM departamento WHERE Nombre = %s", (nombre_departamento,))
        existing_department = cursor.fetchone()

        if existing_department:
            cursor.close()
            # Obtener la lista de departamentos antes del error
            cursor = db.cursor(dictionary=True)
            cursor.execute("SELECT * FROM departamento WHERE estado='Activo'")
            departamentos = cursor.fetchall()
            cursor.close()

            # Ya existe un departamento con el mismo nombre, mostrar alerta
            return render_template("Departamentos.html", error="Ya hay un departamento con este nombre.", departamentos=departamentos)

        if nombre_departamento:
            cursor.execute(
                "INSERT INTO departamento (Nombre, Fecha_Creacion) VALUES (%s, %s)", (nombre_departamento, Fecha_Creacion))
            db.commit()
            cursor.execute("SELECT LAST_INSERT_ID() as id")
            departamento_id = cursor.fetchone()[0]

            # Solución a bugg
            cursor.execute(
                "INSERT INTO cargos (Nombre, Salario_Base, Departamento_ID, Riesgo_LV) VALUES (%s, %s, %s, %s)",
                ("Administrador", 120000, departamento_id, "III")
            )
            db.commit()

            cursor.close()

        # Redirige a la función listar_departamentos después de agregar (incluso si hay un error)
        return redirect(url_for("listar_departamentos"))

    return render_template("Departamentos.html", departamentos=departamentos)


@app.route("/Editar_departamento/<int:id_departamento>", methods=["GET", "POST"])
def Editar_departamento(id_departamento):
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM departamento WHERE ID = %s",
                   (id_departamento,))
    departamento = cursor.fetchone()

    if departamento is None:
        return "Departamento no encontrado."

    departamentos = []  # Inicializar la variable departamentos

    if request.method == "POST":
        nuevo_nombre = request.form["nombreEditar"]
        nueva_fecha = request.form["FechaEditar"]

        cursor = db.cursor()
        cursor.execute("SELECT ID FROM departamento WHERE Nombre = %s AND ID != %s",
                       (nuevo_nombre, id_departamento,))
        existing_department = cursor.fetchone()

        if existing_department:
            cursor.close()
            # Obtener la lista de departamentos antes del error
            cursor = db.cursor(dictionary=True)
            cursor.execute("SELECT * FROM departamento WHERE estado='Activo'")
            departamentos = cursor.fetchall()
            cursor.close()

            # Ya existe un departamento con el mismo nombre, mostrar alerta
            return render_template("Departamentos.html", error="Ya hay un departamento con este nombre.", departamentos=departamentos)

        if nuevo_nombre:
            cursor.execute("UPDATE departamento SET Nombre = %s, Fecha_Creacion = %s WHERE ID = %s",
                           (nuevo_nombre, nueva_fecha, id_departamento,))
            db.commit()
            cursor.close()

            return redirect(url_for("listar_departamentos"))

    return render_template("Departamentos.html", departamento=departamento, departamentos=departamentos)


@app.route("/cargos/<int:departamento_id>")
def obtener_cargos(departamento_id):
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        "SELECT * FROM Cargos WHERE Departamento_ID = %s AND Estado = 'activo'", (departamento_id,))
    cargos = cursor.fetchall()
    cursor.close()

    try:
        # Convertir a JSON utilizando json.dumps
        cargos_json = json.dumps(cargos)
        return cargos_json
    except Exception as e:
        print(f"Error al convertir a JSON: {e}")
        return jsonify({"error": "Error al convertir a JSON"}), 500


@app.route("/agregar_cargos", methods=["GET", "POST"])
def agregar_cargos():

    if "username" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":

        nombre_cargo = request.form["nombre_cargo"]
        salario_base = request.form["salario_base"]
        Departamento_ID = request.form["Departamento_id"]
        nivel_riesgo = request.form["Riesgo_LV"]
        cursor = db.cursor()
        cursor.execute("INSERT INTO Cargos (Nombre, Departamento_id, Salario_Base, Riesgo_LV) VALUES (%s, %s, %s, %s)",
                       (nombre_cargo, Departamento_ID, salario_base, nivel_riesgo))
        db.commit()
        cursor.close()
        return redirect(url_for("cargos"))

    return render_template("cargos.html")


@app.route("/editar_cargo", methods=["POST"])
def editar_cargo():
    if "username" not in session:
        return redirect(url_for("login"))

    if request.method == "POST":
        cargo_id = request.form["cargo_id_edit"]
        nombre_cargo = request.form["nombre_cargo_edit"]
        salario_base = request.form["salario_base_edit"]
        nivel_riesgo = request.form["nivel_riesgo_edit"]

        if nombre_cargo and salario_base and nivel_riesgo:
            cursor = db.cursor()
            update_query = (
                "UPDATE cargos SET Nombre = %s, Salario_Base = %s, Riesgo_LV = %s WHERE ID = %s"
            )
            cursor.execute(update_query, (nombre_cargo,
                           salario_base, nivel_riesgo, cargo_id))
            db.commit()
            cursor.close()

            return redirect(url_for("cargos"))
        else:
            return redirect(url_for("cargos"))


@app.route("/desactivar_cargo", methods=['POST'])
def desactivar_cargo():
    cargo_id = request.form.get('cargo_id')

    cursor = db.cursor()
    try:
        # Actualizar el estado del cargo a 'Inactivo'
        cursor.execute(
            "UPDATE Cargos SET Estado = 'Inactivo' WHERE ID = %s", (cargo_id,))
        db.commit()
        # Redirige a la página de cargos después de la desactivación
        return redirect('/cargos')
    except Exception as e:
        print(f"Error al desactivar el cargo: {e}")
        db.rollback()
        return jsonify({"error": "Error al desactivar el cargo"}), 500
    finally:
        cursor.close()


@app.route("/reactivar_cargo", methods=['POST'])
def reactivar_cargo():
    cargo_id = request.form.get('cargo_id')

    cursor = db.cursor()
    try:
        # Actualizar el estado del cargo a 'Activo'
        cursor.execute(
            "UPDATE Cargos SET Estado = 'Activo' WHERE ID = %s", (cargo_id,))
        db.commit()
        # Redirige a la página de cargos después de la reactivación
        return redirect('/cargos')
    except Exception as e:
        print(f"Error al reactivar el cargo: {e}")
        db.rollback()
        return jsonify({"error": "Error al reactivar el cargo"}), 500
    finally:
        cursor.close()


@app.route("/agregar", methods=["GET", "POST"])
def agregar():
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)

    if request.method == "POST":
        nombres = request.form["nombres"]
        nro_documento = request.form["nro_documento"]
        apellido_paterno = request.form["apellido_paterno"]
        apellido_materno = request.form["apellido_materno"]
        fecha_nacimiento = request.form["fecha_nacimiento"]
        direccion = request.form["direccion"]
        barrio = request.form["barrio"]
        telefono = request.form["telefono"]
        email = request.form["email"]
        departamento_id = int(request.form["departamento"])
        sueldo = int(request.form["sueldo"])
        Tipo_Sangre = request.form["Tipo_Sangre"]
        # Agregado para obtener el cargo seleccionado
        cargo_id = int(request.form["cargo"])

        query = "INSERT INTO empleados (Nombres, Nro_documento, Apellido_Paterno, Apellido_Materno, Fecha_Nacimiento, Direccion, Barrio, Telefono, Email, Departamento, Tipo_Sangre, Sueldo, Cargo_ID) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        values = (nombres, nro_documento, apellido_paterno, apellido_materno, fecha_nacimiento,
                  direccion, barrio, telefono, email, departamento_id, Tipo_Sangre, sueldo, cargo_id)
        cursor.execute(query, values)
        db.commit()
        cursor.close()

        return redirect(url_for("trabajadores"))

    cursor.execute("SELECT * FROM departamento")
    departamentos = cursor.fetchall()
    cursor.close()

    return render_template("agregar.html", departamentos=departamentos)


@app.route("/editar/<int:id_empleado>", methods=["GET", "POST"])
def editar(id_empleado):
    if "username" not in session:
        return redirect(url_for("login"))

    # Configuramos el cursor para devolver diccionarios
    cursor = db.cursor(dictionary=True)

    if request.method == "POST":
        nombres = request.form["nombres"]
        apellido_paterno = request.form["apellido_paterno"]
        apellido_materno = request.form["apellido_materno"]
        fecha_nacimiento = request.form["fecha_nacimiento"]
        direccion = request.form["direccion"]
        barrio = request.form["barrio"]
        telefono = request.form["telefono"]
        email = request.form["email"]
        departamento = int(request.form["departamento"])
        # Asegúrate de convertirlo a float si es un número decimal
        sueldo = float(request.form["sueldo"])
        cargo = int(request.form["cargo"])  # Nuevo campo para el cargo

        query = "UPDATE empleados SET Nombres = %s, Apellido_Paterno = %s, Apellido_Materno = %s, Fecha_Nacimiento = %s, Direccion = %s, Barrio = %s, Telefono = %s, Email = %s, Departamento = %s, Sueldo = %s, Cargo_ID = %s WHERE ID = %s"
        values = (nombres, apellido_paterno, apellido_materno, fecha_nacimiento,
                  direccion, barrio, telefono, email, departamento, sueldo, cargo, id_empleado)
        cursor.execute(query, values)
        db.commit()
        cursor.close()

        return redirect(url_for("trabajadores"))

    cursor.execute("SELECT * FROM empleados WHERE ID = %s", (id_empleado,))
    empleado = cursor.fetchone()

    # Obtener el ID del departamento del empleado
    departamento_empleado = empleado['Departamento']

    # Obtener los cargos del departamento del empleado
    cursor.execute(
        "SELECT * FROM cargos WHERE Departamento_ID = %s", (departamento_empleado,))
    cargos = cursor.fetchall()
    cursor.execute("SELECT * FROM departamento")
    departamentos = cursor.fetchall()
    cursor.close()

    return render_template("editar.html", empleado=empleado, departamentos=departamentos, cargos=cargos)


@app.route("/desactivar/<int:id>")
def desactivar(id):
    if "username" in session:
        cursor = db.cursor()
        cursor.execute(
            "UPDATE empleados SET Estado = 'Inactivo' WHERE ID = %s", (id,))
        db.commit()
        cursor.close()
        return redirect(url_for("inactivos"))
    else:
        return redirect(url_for("login"))


@app.route("/inactivos")
def inactivos():
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM empleados WHERE Estado = 'Inactivo'")
    empleados_inactivos = cursor.fetchall()
    cursor.close()

    return render_template("inactivos.html", empleados_inactivos=empleados_inactivos)


@app.route("/reactivar/<int:id>")
def reactivar_empleado(id):
    if "username" in session:
        cursor = db.cursor()
        cursor.execute(
            "UPDATE empleados SET Estado = 'Activo' WHERE ID = %s", (id,))
        db.commit()
        cursor.close()
        return redirect(url_for("inactivos"))
    else:
        return redirect(url_for("login"))


@app.route("/logout")
def logout():
    session.pop("username", None)
    return redirect(url_for("login"))


# Configura la localización para formatear números con comas
locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')


def format_number(number):
    return locale.format_string("%d", number, grouping=True)


@app.route("/Horario", methods=["GET", "POST"])
def Horario():
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM empleados WHERE Estado = 'activo'")
    empleados = cursor.fetchall()

    if not empleados:
         return render_template(
            "Horario.html",
            empleados=empleados,
            error="true"
        )

    if request.method == "POST":

        empleado_id = request.form["empleado"]
        dias = int(request.form["dias"])
        fecha_inicio = request.form["fecha"]
        cursor.execute(
            "SELECT Sueldo, Nombres, Apellido_Materno, Apellido_Paterno, Nro_documento FROM empleados WHERE ID = %s", (empleado_id,))
        empleado = cursor.fetchone()
        empleado_sueldo = empleado['Sueldo']
        empleado_nombre_completo = f"{empleado['Nombres']} {empleado['Apellido_Paterno']} {empleado['Apellido_Materno']}"
        empleado_Identificacion = empleado['Nro_documento']
        horas_extra_diurnas = int(request.form["horas_extra_diurnas"]) if request.form["horas_extra_diurnas"] else 0
        horas_extra_nocturnas = int(request.form["horas_extra_nocturnas"]) if request.form["horas_extra_nocturnas"] else 0
        horas_dominicales = int(request.form["horas_dominicales"]) if request.form["horas_dominicales"] else 0
        horas_extra_dominicales = int(request.form["horas_extra_dominicales"]) if request.form["horas_extra_dominicales"] else 0
        horas_extra_dom_nocturnas = int(request.form["horas_extra_dom_nocturnas"]) if request.form["horas_extra_dom_nocturnas"] else 0
        Comisiones = int(request.form["Comisiones"]) if request.form["Comisiones"] else 0
        Otros = int(request.form["Otros"]) if request.form["Otros"] else 0

        fecha_actual = datetime.now()
        empleado_sueldo = empleado_sueldo
        salario_base_diario = empleado_sueldo / 26

        empleado_sueldo = salario_base_diario * dias

        horas_extra_diurnas_total = (
            ((empleado_sueldo * 1.25) / 120) * horas_extra_diurnas)
        horas_extra_nocturnas_total = (
            ((empleado_sueldo * 1.75) / 120) * horas_extra_nocturnas)
        horas_dominicales_total = (
            ((empleado_sueldo * 1.75) / 120) * horas_dominicales)
        horas_extra_dominicales_total = (
            ((empleado_sueldo * 2) / 120) * horas_extra_dominicales)
        horas_extra_dom_nocturnas_total = (
            ((empleado_sueldo * 2.5) / 120) * horas_extra_dom_nocturnas)

        cursor = db.cursor()

        horas_extra_total = horas_extra_diurnas_total + horas_extra_nocturnas_total + \
            horas_extra_dominicales_total + horas_extra_dom_nocturnas_total
        salud = (empleado_sueldo + horas_extra_total) * 0.04
        pension = (empleado_sueldo + horas_extra_total) * 0.04

        if (empleado_sueldo > 1160000*2):
            total_devengado = empleado_sueldo + horas_extra_total
        else:
            total_devengado = empleado_sueldo + 140606 + horas_extra_total

        total_deducido = salud + pension
        neto_pagado = total_devengado - total_deducido

        total_deducido = salud + pension
        fecha_actual = datetime.now()

        cursor.execute("INSERT INTO deduccion (Empleado_ID, Comisiones, Salud, Pension, Fecha, Otros) VALUES (%s, %s, %s, %s, %s, %s)",
                   (empleado_id, Comisiones, salud, pension, fecha_actual, Otros))
        cursor.execute("INSERT INTO sueldo (Empleado_ID, Total_Dev, Total_Ded, Salario_Neto, Fecha, Fecha_Inicio) VALUES (%s, %s, %s, %s, %s, %s)",
                   (empleado_id, total_devengado, total_deducido, neto_pagado, fecha_actual, fecha_inicio))
        cursor.execute("INSERT INTO horario (HorExtDiurna, HorExtNocturna, HorExtDominical, HorExtDomNoct, empleado_ID, Fecha) VALUES (%s, %s, %s, %s, %s, %s)",
                   (horas_extra_diurnas_total, horas_extra_nocturnas_total, horas_extra_dominicales_total, horas_extra_dom_nocturnas_total, empleado_id, fecha_actual))

        db.commit()
        cursor.close()

        return render_template(
            "Horario.html",
            empleados=empleados,
            sucess="true"
        )

    return render_template(
        "Horario.html",
        empleados=empleados
    )


def generar_pdf(empleado_nombre, horas_extra_total, total_devengado, total_deducido, neto_pagado, fecha_actual):
    # Obtener la ruta de la carpeta de descargas
    downloads_folder = os.path.expanduser("~") + "/Downloads"

    # Limpieza del nombre del empleado
    empleado_nombre_limpio = re.sub(r'\W+', '', empleado_nombre)

    # Formateo de la fecha para que sea adecuada como parte del nombre del archivo
    fecha_str_limpio = fecha_actual.strftime("%Y%m%d_%H%M%S")

    # Construir la ruta completa del PDF
    pdf_path = os.path.join(
        downloads_folder, f"boletin_pago_{empleado_nombre_limpio}_{fecha_str_limpio}.pdf")

    # Crear el objeto Canvas
    c = canvas.Canvas(pdf_path)

    # Configuración del PDF
    width, height = 400, 600
    c.setPageSize((width, height))

    # Contenido del PDF
    c.drawString(100, height - 50, "Boletín de Pago")
    c.drawString(100, height - 80, f"Empleado: {empleado_nombre}")
    c.drawString(100, height - 110, f"Total Devengado: {total_devengado}")
    c.drawString(100, height - 140, f"Total Deducido: {total_deducido}")
    c.drawString(100, height - 170, f"Salario Neto: {neto_pagado}")

    # Guardar el PDF
    c.save()


@app.route('/reportes')
def reportes():
    # Datos ficticios de ganancias mensuales
    ganancias_mensuales = [
        {'año': '2018', 'ganancia': 1105000},
        {'año': '2019', 'ganancia': 1160000},
        {'año': '2020', 'ganancia': 900000},
        {'año': '2021', 'ganancia': 1000000},
        {'año': '2022', 'ganancia': 2200300},
        # Agrega más datos para otros meses
    ]

    # Convierte los datos en un DataFrame de pandas
    df_ganancias = pd.DataFrame(ganancias_mensuales)

    # Lógica para obtener la cantidad de empleados activos e inactivos
    cursor = db.cursor()
    cursor.execute("SELECT estado, COUNT(*) FROM empleados GROUP BY estado")
    estadisticas_estado = cursor.fetchall()

    # Lógica para obtener estadísticas de empleados por departamento con nombres de departamentos
    cursor.execute("SELECT d.nombre AS departamento, COUNT(*) FROM empleados AS e "
                   "INNER JOIN departamento AS d ON e.departamento = d.id "
                   "GROUP BY e.departamento")
    estadisticas_departamentos = cursor.fetchall()

    # Lógica para calcular el pago total a todos los empleados
    cursor.execute("SELECT SUM(sueldo) FROM empleados")
    pago_total = cursor.fetchone()[0]

    # Crear un gráfico de barras para las estadísticas de departamentos
    df_departamentos = pd.DataFrame(estadisticas_departamentos, columns=[
                                    'Departamento', 'Cantidad'])

    fig_departamentos = px.bar(
        df_departamentos,
        x='Departamento',
        y='Cantidad',
        title='Cantidad de Empleados por Departamento'
    )

    # Crear un gráfico de barras para las estadísticas de empleados activos e inactivos
    df_estado = pd.DataFrame(estadisticas_estado, columns=[
                             'Estado', 'Cantidad'])

    fig_estado = px.bar(
        df_estado,
        x='Estado',
        y='Cantidad',
        title='Empleados Activos vs. Inactivos',
        labels={'Estado': 'Estado de Empleados',
                'Cantidad': 'Cantidad de Empleados'}
    )

    # Crea un gráfico de líneas para las ganancias mensuales
    fig_ganancias = px.line(
        df_ganancias,
        x='año',
        y='ganancia',
        labels={'año': 'Año', 'ganancia': 'Ganancias'}
    )

    # Convierte los gráficos Plotly en HTML
    grafico_departamentos_html = fig_departamentos.to_html(full_html=False)
    grafico_estado_html = fig_estado.to_html(full_html=False)
    grafico_ganancias_html = fig_ganancias.to_html(full_html=False)

    # Texto que muestra el pago total a todos los empleados
    texto_pago_total = f"Pago total general correspondiente a todos los empleados: ${pago_total}"

    return render_template(
        'reportes.html',
        grafico_departamentos_html=grafico_departamentos_html,
        grafico_estado_html=grafico_estado_html,
        grafico_ganancias_html=grafico_ganancias_html,
        texto_pago_total=texto_pago_total  # Agrega esta línea
    )


@app.route("/generar_reporte")
def generar_reporte():
    if "username" not in session:
        return redirect(url_for("login"))

    cursor = db.cursor(dictionary=True)

    # Obtener datos de todos los empleados activos con los nombres de los departamentos
    cursor.execute("SELECT e.*, h.*, d.*, s.* FROM empleados e LEFT JOIN horario h ON e.ID = h.empleado_ID LEFT JOIN deduccion d ON e.ID = d.Empleado_ID LEFT JOIN sueldo s ON e.ID = s.Empleado_ID WHERE e.Estado = 'Activo'")

    empleados_activos = cursor.fetchall()

    cursor.close()

    # Obtener la fecha y hora actual
    now = datetime.now()

    # Formatear la fecha y hora sin los dos puntos
    formatted_time = now.strftime("%Y-%m-%d %H-%M-%S")

    # Crear el título de la hoja sin caracteres no permitidos
    sheet_title = f"Reporte {formatted_time}"

    # Crear un nuevo libro de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_title

    # Agregar encabezado en la fila 2
    encabezado = ['Nombre', 'Nro de documento', 'Apellido Paterno', 'Apellido Materno', 'Edad', 'Dirección', 'Teléfono', 'Email', 'Departamento', 'Salario en Bruto', '', 'Horas Diurnas',
                  'Horas Extra Diurnas', 'Hora Nocturna', 'Hora Extra nocturna', 'Hora dominical', 'Hora extra dominical', 'Hora extra Dominical N.', '', 'Salud', 'Pensión', 'Salario Neto']
    for col_idx, encabezado_text in enumerate(encabezado, 1):
        sheet.cell(row=2, column=col_idx, value=encabezado_text)
        # Agregar color de fondo al encabezado
        sheet.cell(row=2, column=col_idx).fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Ajustar el ancho de la columna según el contenido más largo del encabezado
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        column_width = max(len(encabezado_text), len(
            "Apellido Paterno")) + 6  # Agregar margen adicional
        sheet.column_dimensions[column_letter].width = column_width

    # Agregar datos al reporte a partir de la fila 3
    for row_idx, empleado in enumerate(empleados_activos, start=3):
        sheet.cell(row=row_idx, column=1, value=empleado['Nro_documento'])
        sheet.cell(row=row_idx, column=2, value=empleado['Nombres'])
        sheet.cell(row=row_idx, column=3, value=empleado['Apellido_Paterno'])
        sheet.cell(row=row_idx, column=4, value=empleado['Apellido_Materno'])
        sheet.cell(row=row_idx, column=5, value=empleado['Edad'])
        sheet.cell(row=row_idx, column=6, value=empleado['Direccion'])
        sheet.cell(row=row_idx, column=7, value=empleado['Telefono'])
        sheet.cell(row=row_idx, column=8, value=empleado['Email'])
        sheet.cell(row=row_idx, column=9, value=empleado['Departamento'])
        sheet.cell(row=row_idx, column=10, value=empleado['Total_Dev'])
        sheet.cell(row=row_idx, column=11, value=empleado['HorDiurna'])
        sheet.cell(row=row_idx, column=12, value=empleado['HorExtDiurna'])
        sheet.cell(row=row_idx, column=13, value=empleado['HorNocturna'])
        sheet.cell(row=row_idx, column=14, value=empleado['HorExtNocturna'])
        sheet.cell(row=row_idx, column=15, value=empleado['HorDominical'])
        sheet.cell(row=row_idx, column=16, value=empleado['HorExtDominical'])
        sheet.cell(row=row_idx, column=17, value=empleado['HorExtDomNoct'])
        sheet.cell(row=row_idx, column=18, value=empleado['Salud'])
        sheet.cell(row=row_idx, column=19, value=empleado['Pension'])
        sheet.cell(row=row_idx, column=20, value=empleado['Salario_Neto'])

        # Aplicar bordes a las celdas con contenido
        for col_idx in range(1, 22):
            celda = sheet.cell(row=row_idx, column=col_idx)
            if celda.value:
                borde = Border(left=Side(style='thin', color='000000'),
                               right=Side(style='thin', color='000000'),
                               top=Side(style='thin', color='000000'),
                               bottom=Side(style='thin', color='000000'))
                celda.border = borde

    # Crear una respuesta para descargar el archivo
    response = Response(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_empleados.xlsx'
    workbook.save(response.stream)
    return response


if __name__ == "__main__":
    app.run(debug=True)
